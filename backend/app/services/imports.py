import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session

# Teto de upload para evitar consumo excessivo de memoria (arquivo lido inteiro em RAM).
MAX_IMPORT_BYTES = 5 * 1024 * 1024  # 5 MB

from app.models.student import Student
from app.models.user import User
from app.schemas.imports import ImportErrorRow, StudentImportResult
from app.schemas.student import StudentCreate
from app.services.audit import record_audit

HEADER_ALIASES = {
    "nome": "name",
    "name": "name",
    "telefone": "phone",
    "phone": "phone",
    "email": "email",
    "e-mail": "email",
    "cpf": "cpf",
    "data_nascimento": "birth_date",
    "nascimento": "birth_date",
    "birth_date": "birth_date",
    "plano": "plan",
    "plan": "plan",
    "mensalidade": "monthly_fee",
    "valor_mensalidade": "monthly_fee",
    "monthly_fee": "monthly_fee",
    "vencimento": "due_day",
    "dia_vencimento": "due_day",
    "due_day": "due_day",
    "status": "status",
    "observacoes": "notes",
    "observações": "notes",
    "notes": "notes",
}


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def normalize_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def parse_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    return Decimal(text)


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("Data invalida. Use AAAA-MM-DD ou DD/MM/AAAA.")


def map_row(raw: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for header, value in raw.items():
        field = HEADER_ALIASES.get(normalize_header(header))
        if field:
            mapped[field] = normalize_value(value)

    mapped["email"] = mapped.get("email") or None
    mapped["cpf"] = mapped.get("cpf") or None
    mapped["birth_date"] = parse_date(mapped.get("birth_date"))
    mapped["monthly_fee"] = parse_decimal(mapped.get("monthly_fee"))
    mapped["due_day"] = int(mapped.get("due_day") or 1)
    mapped["status"] = str(mapped.get("status") or "ATIVO").upper()
    mapped["notes"] = mapped.get("notes") or None
    return mapped


async def read_spreadsheet_rows(file: UploadFile) -> list[dict[str, Any]]:
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo vazio.")
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Arquivo muito grande. Tamanho maximo: {MAX_IMPORT_BYTES // (1024 * 1024)} MB.",
        )

    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        sample = text[:2048]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        return list(csv.DictReader(StringIO(text), delimiter=delimiter))

    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(header or "").strip() for header in rows[0]]
        return [
            {headers[index]: cell for index, cell in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(cell not in (None, "") for cell in row)
        ]

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie um arquivo .csv ou .xlsx.")


async def import_students_from_file(db: Session, file: UploadFile, current_user: User) -> StudentImportResult:
    rows = await read_spreadsheet_rows(file)
    errors: list[ImportErrorRow] = []
    imported = 0
    seen_cpfs: set[str] = set()
    seen_emails: set[str] = set()

    for index, row in enumerate(rows, start=2):
        try:
            payload = StudentCreate(**map_row(row))
        except (ValueError, ValidationError, InvalidOperation, ArithmeticError) as exc:
            message = str(exc).splitlines()[0] if str(exc).strip() else "Linha invalida."
            errors.append(ImportErrorRow(row=index, message=message))
            continue

        # Evita duplicar CPF/e-mail contra o banco e dentro do proprio arquivo.
        email_key = payload.email.lower() if payload.email else None
        conflict: str | None = None
        if payload.cpf and (
            payload.cpf in seen_cpfs
            or db.scalar(select(Student.id).where(Student.cpf == payload.cpf)) is not None
        ):
            conflict = f"CPF {payload.cpf} ja cadastrado."
        elif email_key and (
            email_key in seen_emails
            or db.scalar(select(Student.id).where(func.lower(Student.email) == email_key)) is not None
        ):
            conflict = f"E-mail {payload.email} ja cadastrado."
        if conflict:
            errors.append(ImportErrorRow(row=index, message=conflict))
            continue

        if payload.cpf:
            seen_cpfs.add(payload.cpf)
        if email_key:
            seen_emails.add(email_key)
        student = Student(**payload.model_dump())
        db.add(student)
        imported += 1

    record_audit(
        db,
        current_user,
        entity_type="STUDENT",
        action="IMPORT",
        summary=f"Importacao de alunos: {imported} importados, {len(errors)} erros.",
        after={"imported": imported, "errors": [error.model_dump() for error in errors]},
    )
    db.commit()
    return StudentImportResult(imported=imported, skipped=len(errors), errors=errors)
